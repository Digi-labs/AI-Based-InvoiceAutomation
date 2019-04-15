#Usage: python app.py
import os
 
from flask import Flask, render_template, request, redirect, url_for
from werkzeug import secure_filename
from flask import Flask, Response
from flask import send_file

import argparse
import imutils
import PyPDF2
from PyPDF2 import PdfFileWriter, PdfFileReader
from openpyxl import Workbook

import time

import uuid
import base64


UPLOAD_FOLDER = 'uploads'


def get_as_base64(url):
    return base64.b64encode(requests.get(url).content)

def predict(file):
    # Set input and output file names 
    
    output_file = 'digiinvoice.xlsx'

    # Open & Read PDF
    pdf_file = open(file,'rb')
    input_pdf = PyPDF2.PdfFileReader(pdf_file)

    # Declare Headings of PDF file
    main_list = ['From',
     'To',
     'Invoice Number',
     'Order Number',
     'Invoice Date',
     'Due Date',
     'Total Due',
     'Quantity',
     'Service',
     'Rate',
     'Adjust',
     'Sub Total','#@%^&*']


    # New workbook in openpyxl 
    wb = Workbook()
    ws = wb.active

    # Write Headings into excel file
    row_num=1
    column_num=1
    for i in range(len(main_list)-1):
        field = main_list[i]
        ws.cell(row=row_num, column=column_num, value=field)
        column_num += 1

    # Count total page number of PDF file
    total_pages = input_pdf.getNumPages()

    # Extract data from PDF and Write it into excel file
    row_num = 2
    for i in range(total_pages):
        page = input_pdf.getPage(i)
        page_content = page.extractText()
        column_num = 1
        for i in range(len(main_list)-1):
                field = main_list[i]
                next_field = main_list[i+1]
                # Find position of fields from extracted text of PDF file
                field_pos = page_content.find(field)
                next_field_pos = page_content.find(next_field)
                # Find position of field values from extracted text of PDF file
                field_value_start_pos = field_pos+len(field)
                field_value_end_pos = next_field_pos
                # Extract field values
                field_value = page_content[field_value_start_pos:field_value_end_pos]
                # Write field values into Excel
                ws.cell(row = row_num, column = column_num, value = field_value)
                column_num += 1
        row_num += 1

    pdf_file.close()

    # Save excel file
    wb.save(output_file)

    return output_file




def my_random_string(string_length=10):
    """Returns a random string of length string_length."""
    random = str(uuid.uuid4()) # Convert UUID format to a Python string.
    random = random.upper() # Make all characters uppercase.
    random = random.replace("-","") # Remove the UUID '-'.
    return random[0:string_length] # Return the random string.



app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route("/")
def template_test():
    return render_template('template.html', label='')

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        import time
        start_time = time.time()
        file = request.files['file']

     
        filename = secure_filename(file.filename)

        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        result = predict(file_path)
            			
            
        print(file_path)
        filename = my_random_string(6) + filename

        os.rename(file_path, os.path.join(app.config['UPLOAD_FOLDER'], filename))
        print("--- %s seconds ---" % str (time.time() - start_time))
        return render_template('template.html',label="Download File" )

from flask import send_from_directory

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'],
                               filename)

from werkzeug import SharedDataMiddleware
app.add_url_rule('/uploads/<filename>', 'uploaded_file',
                 build_only=True)
app.wsgi_app = SharedDataMiddleware(app.wsgi_app, {
    '/uploads':  app.config['UPLOAD_FOLDER']
})
@app.route('/getPlotCSV') # this is a job for GET, not POST
def plot_csv():
    return send_file('digiinvoice.xlsx',
                     mimetype='text/xlsx',
                     attachment_filename='digiinvoice.xlsx',
                     as_attachment=True)

if __name__ == "__main__":
    app.debug=False
    app.run(host='0.0.0.0', port=3000)